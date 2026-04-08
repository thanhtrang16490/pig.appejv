#!/usr/bin/env python3
"""
Automated Database Deployment and Data Import Script
- Runs migration SQL against Supabase
- Imports data from Excel file
- Verifies the import
"""

import os
import sys
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Optional
import subprocess

# Check and install required packages
def ensure_packages():
    """Ensure required packages are installed."""
    required = ['openpyxl', 'supabase', 'httpx']
    for package in required:
        try:
            __import__(package.replace('-', '_'))
        except ImportError:
            print(f"Installing {package}...")
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', package, '-q'])

ensure_packages()

import httpx
from openpyxl import load_workbook
from supabase import create_client, Client

# Supabase configuration
SUPABASE_URL = "https://svlgjtsjcsksqajtwepc.supabase.co"
SUPABASE_SERVICE_ROLE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InN2bGdqdHNqY3Nrc3FhanR3ZXBjIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NTYxMzk0MiwiZXhwIjoyMDkxMTg5OTQyfQ.Gce32hBhbrw-O--z2Wmu2opO9XgSXUasEfpCQdtfgHc"
SUPABASE_PROJECT_REF = "svlgjtsjcsksqajtwepc"

# Migration SQL
MIGRATION_SQL = """
-- Initial Schema Migration for Pig Farm Management System

-- ============================================
-- ENUMS
-- ============================================

DO $$
BEGIN
    IF NOT EXISTS (SELECT 1 FROM pg_type WHERE typname = 'product_type') THEN
        CREATE TYPE product_type AS ENUM ('heo_thit', 'heo_con');
    END IF;
    IF NOT EXISTS (SELECT 1 FROM pg_type WHERE typname = 'payment_method') THEN
        CREATE TYPE payment_method AS ENUM ('cash', 'bank', 'company');
    END IF;
END
$$;

-- ============================================
-- TABLES
-- ============================================

-- Farms table
CREATE TABLE IF NOT EXISTS farms (
    id uuid PRIMARY KEY DEFAULT gen_random_uuid(),
    name text NOT NULL,
    group_name text,
    location text,
    created_at timestamptz DEFAULT now()
);

-- Customers table
CREATE TABLE IF NOT EXISTS customers (
    id uuid PRIMARY KEY DEFAULT gen_random_uuid(),
    name text NOT NULL,
    address text,
    phone text,
    notes text,
    created_at timestamptz DEFAULT now()
);

-- Transactions table
CREATE TABLE IF NOT EXISTS transactions (
    id uuid PRIMARY KEY DEFAULT gen_random_uuid(),
    transaction_date date NOT NULL,
    farm_id uuid NOT NULL REFERENCES farms(id) ON DELETE RESTRICT,
    customer_id uuid NOT NULL REFERENCES customers(id) ON DELETE RESTRICT,
    product_type product_type NOT NULL,
    quantity integer NOT NULL,
    total_weight_kg numeric NOT NULL DEFAULT 0,
    avg_weight_kg numeric DEFAULT 0,
    unit_price numeric NOT NULL DEFAULT 0,
    revenue_by_count numeric DEFAULT 0,
    excess_weight_kg numeric DEFAULT 0,
    excess_price_per_kg numeric DEFAULT 0,
    excess_revenue numeric DEFAULT 0,
    total_invoice numeric DEFAULT 0,
    payment_cash numeric DEFAULT 0,
    payment_bank numeric DEFAULT 0,
    payment_company numeric DEFAULT 0,
    total_paid numeric DEFAULT 0,
    outstanding_debt numeric DEFAULT 0,
    surplus numeric DEFAULT 0,
    notes text,
    created_at timestamptz DEFAULT now(),
    updated_at timestamptz DEFAULT now()
);

-- Payments table
CREATE TABLE IF NOT EXISTS payments (
    id uuid PRIMARY KEY DEFAULT gen_random_uuid(),
    transaction_id uuid REFERENCES transactions(id) ON DELETE SET NULL,
    customer_id uuid NOT NULL REFERENCES customers(id) ON DELETE RESTRICT,
    payment_date date NOT NULL,
    amount numeric NOT NULL,
    method payment_method NOT NULL,
    notes text,
    created_at timestamptz DEFAULT now()
);

-- ============================================
-- INDEXES
-- ============================================

-- Transactions indexes
CREATE INDEX IF NOT EXISTS idx_transactions_date ON transactions(transaction_date);
CREATE INDEX IF NOT EXISTS idx_transactions_farm ON transactions(farm_id);
CREATE INDEX IF NOT EXISTS idx_transactions_customer ON transactions(customer_id);
CREATE INDEX IF NOT EXISTS idx_transactions_product_type ON transactions(product_type);

-- Payments indexes
CREATE INDEX IF NOT EXISTS idx_payments_customer ON payments(customer_id);
CREATE INDEX IF NOT EXISTS idx_payments_transaction ON payments(transaction_id);
CREATE INDEX IF NOT EXISTS idx_payments_date ON payments(payment_date);

-- Customers index for search
CREATE INDEX IF NOT EXISTS idx_customers_name ON customers(name);

-- ============================================
-- VIEWS
-- ============================================

-- Customer summary view
CREATE OR REPLACE VIEW customer_summary AS
SELECT 
    c.id as customer_id,
    c.name as customer_name,
    c.address as customer_address,
    COALESCE(SUM(t.total_invoice), 0) as total_revenue,
    COALESCE(SUM(t.total_paid), 0) as total_paid,
    COALESCE(SUM(t.outstanding_debt), 0) as outstanding_debt,
    COUNT(t.id) as transaction_count
FROM customers c
LEFT JOIN transactions t ON c.id = t.customer_id
GROUP BY c.id, c.name, c.address;

-- Farm summary view
CREATE OR REPLACE VIEW farm_summary AS
SELECT
    f.id as farm_id,
    f.name as farm_name,
    COALESCE(SUM(t.total_invoice), 0) as total_revenue,
    COALESCE(SUM(t.quantity), 0) as total_quantity,
    COALESCE(SUM(t.total_weight_kg), 0) as total_weight_kg,
    COUNT(t.id) as transaction_count
FROM farms f
LEFT JOIN transactions t ON f.id = t.farm_id
GROUP BY f.id, f.name;

-- Daily summary view
CREATE OR REPLACE VIEW daily_summary AS
SELECT
    transaction_date,
    SUM(total_invoice) as total_revenue,
    COUNT(*) as total_transactions,
    SUM(quantity) as total_quantity
FROM transactions
GROUP BY transaction_date
ORDER BY transaction_date DESC;

-- ============================================
-- ROW LEVEL SECURITY
-- ============================================

-- Enable RLS on all tables
ALTER TABLE farms ENABLE ROW LEVEL SECURITY;
ALTER TABLE customers ENABLE ROW LEVEL SECURITY;
ALTER TABLE transactions ENABLE ROW LEVEL SECURITY;
ALTER TABLE payments ENABLE ROW LEVEL SECURITY;

-- Create permissive policies for all tables (internal tool - no auth required)
DO $$
BEGIN
    IF NOT EXISTS (SELECT 1 FROM pg_policies WHERE tablename = 'farms' AND policyname = 'Allow all') THEN
        CREATE POLICY "Allow all" ON farms FOR ALL USING (true) WITH CHECK (true);
    END IF;
    IF NOT EXISTS (SELECT 1 FROM pg_policies WHERE tablename = 'customers' AND policyname = 'Allow all') THEN
        CREATE POLICY "Allow all" ON customers FOR ALL USING (true) WITH CHECK (true);
    END IF;
    IF NOT EXISTS (SELECT 1 FROM pg_policies WHERE tablename = 'transactions' AND policyname = 'Allow all') THEN
        CREATE POLICY "Allow all" ON transactions FOR ALL USING (true) WITH CHECK (true);
    END IF;
    IF NOT EXISTS (SELECT 1 FROM pg_policies WHERE tablename = 'payments' AND policyname = 'Allow all') THEN
        CREATE POLICY "Allow all" ON payments FOR ALL USING (true) WITH CHECK (true);
    END IF;
END
$$;

-- ============================================
-- TRIGGERS
-- ============================================

-- Function to update updated_at column
CREATE OR REPLACE FUNCTION update_updated_at_column()
RETURNS TRIGGER AS $$
BEGIN
    NEW.updated_at = now();
    RETURN NEW;
END;
$$ language 'plpgsql';

-- Trigger for transactions table
DROP TRIGGER IF EXISTS update_transactions_updated_at ON transactions;
CREATE TRIGGER update_transactions_updated_at
    BEFORE UPDATE ON transactions
    FOR EACH ROW EXECUTE FUNCTION update_updated_at_column();
"""

# Excel file path
EXCEL_FILE = "/Users/thanhtrang/Documents/pig.appejv/data/BÁN HÀNG THÁNG 10.2025.xlsx"

# Sheet names
SHEET_HEO_THIT = "Heo thịt_Nhập dữ liệu hàng ngày"
SHEET_HEO_CON = "Heo con_Nhập dữ liệu"
SHEET_DATA = "Data"


def get_supabase_client() -> Client:
    """Create and return Supabase client with service role key."""
    return create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)


def run_migration_via_management_api():
    """Run the database migration using Supabase Management API."""
    print("\n" + "="*60)
    print("STEP 1: Running database migration...")
    print("="*60)
    
    # Use the Supabase Management API to execute SQL
    # First, let's try using the exec_sql function via RPC if it exists
    supabase = get_supabase_client()
    
    try:
        # Try to execute SQL via the Supabase client
        # Split SQL into individual statements and execute them one by one
        statements = [s.strip() for s in MIGRATION_SQL.split(';') if s.strip()]
        
        print("   Creating database objects...")
        
        # Use raw SQL execution via the REST API
        headers = {
            "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
            "Content-Type": "application/json",
            "apikey": SUPABASE_SERVICE_ROLE_KEY
        }
        
        # Try using the SQL execution endpoint
        # The Supabase Management API endpoint for SQL execution
        management_url = f"https://api.supabase.com/v1/projects/{SUPABASE_PROJECT_REF}/database/query"
        
        # For now, let's try a simpler approach - use the Supabase client
        # to check if tables exist, and if not, print the SQL for manual execution
        
        # Check if farms table exists
        try:
            result = supabase.table("farms").select("count", count="exact").limit(1).execute()
            print("   ✅ Tables already exist, skipping migration.")
            return
        except Exception as e:
            if "Could not find the table" in str(e) or "relation" in str(e).lower():
                print("   Tables do not exist. Attempting to create via SQL...")
            else:
                raise e
        
        # Try using the exec_sql RPC function if available
        try:
            result = supabase.rpc('exec_sql', {'sql': MIGRATION_SQL}).execute()
            print("✅ Migration completed successfully via exec_sql!")
            return
        except Exception as e:
            print(f"   exec_sql not available: {e}")
        
        # Try Management API
        try:
            response = httpx.post(
                management_url,
                headers={"Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}"},
                json={"query": MIGRATION_SQL},
                timeout=60.0
            )
            if response.status_code == 200:
                print("✅ Migration completed successfully via Management API!")
                return
            else:
                print(f"   Management API error: {response.status_code} - {response.text}")
        except Exception as e:
            print(f"   Management API not available: {e}")
        
        # If we get here, we need to use the SQL Editor approach
        # But first, let's try using pg_execute if available
        try:
            # Split into smaller chunks and try executing via pg_execute
            # This is a workaround for executing SQL without direct DB access
            print("   Attempting alternative migration methods...")
            
            # Create a temporary function to execute SQL
            create_exec_sql = """
            CREATE OR REPLACE FUNCTION exec_sql(sql text)
            RETURNS void AS $$
            BEGIN
                EXECUTE sql;
            END;
            $$ LANGUAGE plpgsql SECURITY DEFINER;
            """
            
            # Try to create the function using REST API
            rest_url = f"{SUPABASE_URL}/rest/v1/rpc/exec_sql"
            response = httpx.post(
                rest_url,
                headers=headers,
                json={"sql": create_exec_sql},
                timeout=30.0
            )
            
            if response.status_code in [200, 201, 204]:
                print("   Created exec_sql function, retrying migration...")
                result = supabase.rpc('exec_sql', {'sql': MIGRATION_SQL}).execute()
                print("✅ Migration completed successfully!")
                return
        except Exception as e:
            print(f"   Alternative method failed: {e}")
        
        # Final fallback: Print SQL for manual execution
        print("\n" + "="*60)
        print("⚠️  AUTOMATIC MIGRATION NOT POSSIBLE")
        print("="*60)
        print("Please run the following SQL in your Supabase Dashboard:")
        print("1. Go to: https://supabase.com/dashboard/project/svlgjtsjcsksqajtwepc")
        print("2. Navigate to SQL Editor")
        print("3. Create a new query")
        print("4. Paste the SQL below and run it")
        print("="*60)
        print(MIGRATION_SQL)
        print("="*60)
        
        raise Exception("Could not execute migration automatically. Please run the SQL manually in Supabase Dashboard.")
        
    except Exception as e:
        print(f"❌ Migration error: {e}")
        raise


def safe_numeric(value: Any, default: float = 0) -> float:
    """Safely convert a value to float, returning default if None or invalid."""
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            cleaned = value.replace(",", "").replace(" ", "").strip()
            if cleaned == "":
                return default
            return float(cleaned)
        except (ValueError, TypeError):
            return default
    return default


def safe_int(value: Any, default: int = 0) -> int:
    """Safely convert a value to int, returning default if None or invalid."""
    if value is None:
        return default
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        try:
            cleaned = value.replace(",", "").replace(" ", "").strip()
            if cleaned == "":
                return default
            return int(float(cleaned))
        except (ValueError, TypeError):
            return default
    return default


def parse_date(value: Any) -> Optional[datetime]:
    """Parse date from Excel cell value."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]
        for fmt in formats:
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue
    return None


def clean_string(value: Any) -> Optional[str]:
    """Clean and return string value, or None if empty."""
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned if cleaned else None
    return str(value).strip()


def read_customer_data_from_data_sheet(workbook) -> dict:
    """Read customer master data from 'Data' sheet (name -> address mapping)."""
    customer_data = {}
    
    if SHEET_DATA not in workbook.sheetnames:
        print(f"   ⚠️  '{SHEET_DATA}' sheet not found")
        return customer_data
    
    sheet = workbook[SHEET_DATA]
    print(f"\n   Reading customer data from '{SHEET_DATA}' sheet...")
    
    header_row = None
    name_col = None
    address_col = None
    
    for row_idx, row in enumerate(sheet.iter_rows(max_row=20), 1):
        for col_idx, cell in enumerate(row, 1):
            if cell.value and isinstance(cell.value, str):
                cell_text = cell.value.lower().strip()
                if "tên" in cell_text and "khách" in cell_text:
                    header_row = row_idx
                    name_col = col_idx
                elif "địa chỉ" in cell_text or "dia chi" in cell_text:
                    address_col = col_idx
    
    if header_row is None:
        print(f"   ⚠️  Could not find header row in '{SHEET_DATA}' sheet")
        return customer_data
    
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if name_col and len(row) >= name_col:
            name = clean_string(row[name_col - 1])
            if name:
                address = None
                if address_col and len(row) >= address_col:
                    address = clean_string(row[address_col - 1])
                customer_data[name] = address
    
    print(f"   ✅ Found {len(customer_data)} customers in Data sheet")
    return customer_data


def parse_transaction_row(row: tuple, product_type: str, row_num: int, customer_data: dict) -> Optional[dict]:
    """Parse a transaction row from the Excel sheet."""
    try:
        farm_group = clean_string(row[1]) if len(row) > 1 else None
        date_value = row[2] if len(row) > 2 else None
        farm_name = clean_string(row[3]) if len(row) > 3 else None
        customer_name = clean_string(row[4]) if len(row) > 4 else None
        address = clean_string(row[5]) if len(row) > 5 else None
        quantity = safe_int(row[6], 0)
        
        if not customer_name or quantity <= 0:
            return None
        
        transaction_date = parse_date(date_value)
        if transaction_date is None:
            return None
        
        transaction = {
            "transaction_date": transaction_date.strftime("%Y-%m-%d"),
            "farm_name": farm_name or "Unknown Farm",
            "farm_group": farm_group,
            "customer_name": customer_name,
            "address": address or customer_data.get(customer_name),
            "product_type": product_type,
            "quantity": quantity,
            "total_weight_kg": safe_numeric(row[7], 0),
            "avg_weight_kg": safe_numeric(row[8], 0),
            "unit_price": safe_numeric(row[9], 0),
            "revenue_by_count": safe_numeric(row[10], 0),
            "excess_weight_kg": safe_numeric(row[11], 0),
            "excess_price_per_kg": safe_numeric(row[12], 0),
            "excess_revenue": safe_numeric(row[13], 0),
            "total_invoice": safe_numeric(row[14], 0),
            "payment_cash": safe_numeric(row[15], 0),
            "payment_bank": safe_numeric(row[16], 0),
            "payment_company": safe_numeric(row[17], 0),
            "total_paid": safe_numeric(row[18], 0),
            "outstanding_debt": safe_numeric(row[19], 0),
            "surplus": safe_numeric(row[20], 0),
            "notes": clean_string(row[21]) if len(row) > 21 else None,
        }
        
        return transaction
        
    except Exception as e:
        print(f"   Error parsing row {row_num}: {e}")
        return None


def read_transactions_from_sheet(workbook, sheet_name: str, product_type: str, customer_data: dict) -> list:
    """Read all transactions from a sheet."""
    transactions = []
    
    if sheet_name not in workbook.sheetnames:
        print(f"   ⚠️  Sheet '{sheet_name}' not found")
        return transactions
    
    sheet = workbook[sheet_name]
    print(f"\n   Reading transactions from '{sheet_name}'...")
    
    header_row = None
    for row_idx, row in enumerate(sheet.iter_rows(max_row=30), 1):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cell_text = cell.value.lower().strip()
                if any(keyword in cell_text for keyword in ["ngày", "ngay", "thứ", "thu", "date"]):
                    header_row = row_idx
                    break
                if "tên" in cell_text and "khách" in cell_text:
                    header_row = row_idx
                    break
        if header_row:
            break
    
    if header_row is None:
        header_row = 2
    
    data_start_row = header_row + 1
    row_count = 0
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        transaction = parse_transaction_row(row, product_type, row_idx, customer_data)
        if transaction:
            transactions.append(transaction)
            row_count += 1
    
    print(f"   ✅ Parsed {row_count} valid transactions from '{sheet_name}'")
    return transactions


def insert_farms(supabase: Client, farms: dict) -> dict:
    """Insert farms into database and return name -> id mapping."""
    farm_ids = {}
    inserted_count = 0
    
    print(f"\n   Inserting {len(farms)} farms...")
    
    for farm_name, farm_data in farms.items():
        try:
            result = supabase.table("farms").insert(farm_data).execute()
            if result.data:
                farm_ids[farm_name] = result.data[0]["id"]
                inserted_count += 1
        except Exception as e:
            try:
                existing = supabase.table("farms").select("id").eq("name", farm_name).execute()
                if existing.data:
                    farm_ids[farm_name] = existing.data[0]["id"]
            except Exception as e2:
                print(f"   Error handling farm '{farm_name}': {e2}")
    
    print(f"   ✅ Inserted {inserted_count} new farms, total mapped: {len(farm_ids)}")
    return farm_ids


def insert_customers(supabase: Client, customers: dict) -> dict:
    """Insert customers into database and return name -> id mapping."""
    customer_ids = {}
    inserted_count = 0
    
    print(f"\n   Inserting {len(customers)} customers...")
    
    for customer_name, customer_data in customers.items():
        try:
            result = supabase.table("customers").insert(customer_data).execute()
            if result.data:
                customer_ids[customer_name] = result.data[0]["id"]
                inserted_count += 1
        except Exception as e:
            try:
                existing = supabase.table("customers").select("id").eq("name", customer_name).execute()
                if existing.data:
                    customer_ids[customer_name] = existing.data[0]["id"]
            except Exception as e2:
                print(f"   Error handling customer '{customer_name}': {e2}")
    
    print(f"   ✅ Inserted {inserted_count} new customers, total mapped: {len(customer_ids)}")
    return customer_ids


def insert_transactions(supabase: Client, transactions: list, farm_ids: dict, customer_ids: dict) -> int:
    """Insert transactions into database."""
    inserted_count = 0
    error_count = 0
    
    print(f"\n   Inserting {len(transactions)} transactions...")
    
    for idx, txn in enumerate(transactions, 1):
        try:
            farm_name = txn.pop("farm_name")
            customer_name = txn.pop("customer_name")
            address = txn.pop("address", None)
            farm_group = txn.pop("farm_group", None)
            
            farm_id = farm_ids.get(farm_name)
            customer_id = customer_ids.get(customer_name)
            
            if not farm_id:
                error_count += 1
                continue
            
            if not customer_id:
                error_count += 1
                continue
            
            transaction_data = {
                "transaction_date": txn["transaction_date"],
                "farm_id": farm_id,
                "customer_id": customer_id,
                "product_type": txn["product_type"],
                "quantity": txn["quantity"],
                "total_weight_kg": txn["total_weight_kg"],
                "avg_weight_kg": txn["avg_weight_kg"],
                "unit_price": txn["unit_price"],
                "revenue_by_count": txn["revenue_by_count"],
                "excess_weight_kg": txn["excess_weight_kg"],
                "excess_price_per_kg": txn["excess_price_per_kg"],
                "excess_revenue": txn["excess_revenue"],
                "total_invoice": txn["total_invoice"],
                "payment_cash": txn["payment_cash"],
                "payment_bank": txn["payment_bank"],
                "payment_company": txn["payment_company"],
                "total_paid": txn["total_paid"],
                "outstanding_debt": txn["outstanding_debt"],
                "surplus": txn["surplus"],
                "notes": txn["notes"],
            }
            
            result = supabase.table("transactions").insert(transaction_data).execute()
            if result.data:
                inserted_count += 1
            
            if idx % 50 == 0:
                print(f"      Progress: {idx}/{len(transactions)} processed")
                
        except Exception as e:
            error_count += 1
    
    print(f"   ✅ Inserted {inserted_count} transactions, {error_count} errors")
    return inserted_count


def verify_import(supabase: Client):
    """Verify the import by querying record counts."""
    print("\n" + "="*60)
    print("STEP 3: Verification Results")
    print("="*60)
    
    try:
        farms_result = supabase.table("farms").select("*", count="exact").execute()
        farm_count = farms_result.count if hasattr(farms_result, 'count') else len(farms_result.data)
        print(f"   📊 Farms in database: {farm_count}")
        
        customers_result = supabase.table("customers").select("*", count="exact").execute()
        customer_count = customers_result.count if hasattr(customers_result, 'count') else len(customers_result.data)
        print(f"   📊 Customers in database: {customer_count}")
        
        transactions_result = supabase.table("transactions").select("*", count="exact").execute()
        txn_count = transactions_result.count if hasattr(transactions_result, 'count') else len(transactions_result.data)
        print(f"   📊 Transactions in database: {txn_count}")
        
        heo_thit_result = supabase.table("transactions").select("*", count="exact").eq("product_type", "heo_thit").execute()
        heo_con_result = supabase.table("transactions").select("*", count="exact").eq("product_type", "heo_con").execute()
        
        heo_thit_count = heo_thit_result.count if hasattr(heo_thit_result, 'count') else len(heo_thit_result.data)
        heo_con_count = heo_con_result.count if hasattr(heo_con_result, 'count') else len(heo_con_result.data)
        
        print(f"      - Heo thịt (meat hog) transactions: {heo_thit_count}")
        print(f"      - Heo con (piglet) transactions: {heo_con_count}")
        
        if transactions_result.data:
            print("\n   Sample transaction:")
            sample = transactions_result.data[0]
            print(f"      Date: {sample.get('transaction_date')}")
            print(f"      Product: {sample.get('product_type')}")
            print(f"      Quantity: {sample.get('quantity')}")
            print(f"      Total Invoice: {sample.get('total_invoice'):,.0f} VND")
        
        return {
            "farms": farm_count,
            "customers": customer_count,
            "transactions": txn_count,
            "heo_thit": heo_thit_count,
            "heo_con": heo_con_count
        }
        
    except Exception as e:
        print(f"   ❌ Error verifying import: {e}")
        return None


def import_data():
    """Import data from Excel file."""
    print("\n" + "="*60)
    print("STEP 2: Importing data from Excel...")
    print("="*60)
    
    if not os.path.exists(EXCEL_FILE):
        print(f"   ❌ Error: Excel file not found at {EXCEL_FILE}")
        sys.exit(1)
    
    print(f"   Excel file: {EXCEL_FILE}")
    
    try:
        workbook = load_workbook(EXCEL_FILE, data_only=True)
        print(f"   Available sheets: {workbook.sheetnames}")
    except Exception as e:
        print(f"   ❌ Error loading Excel file: {e}")
        sys.exit(1)
    
    customer_data = read_customer_data_from_data_sheet(workbook)
    
    heo_thit_transactions = read_transactions_from_sheet(
        workbook, SHEET_HEO_THIT, "heo_thit", customer_data
    )
    heo_con_transactions = read_transactions_from_sheet(
        workbook, SHEET_HEO_CON, "heo_con", customer_data
    )
    
    all_transactions = heo_thit_transactions + heo_con_transactions
    print(f"\n   Total transactions to import: {len(all_transactions)}")
    
    if not all_transactions:
        print("   No transactions found to import. Exiting.")
        sys.exit(0)
    
    farms = {}
    for txn in all_transactions:
        farm_name = txn["farm_name"]
        if farm_name not in farms:
            farms[farm_name] = {
                "name": farm_name,
                "group_name": txn.get("farm_group"),
                "location": None
            }
    
    customers = {}
    for txn in all_transactions:
        customer_name = txn["customer_name"]
        if customer_name not in customers:
            customers[customer_name] = {
                "name": customer_name,
                "address": txn.get("address") or customer_data.get(customer_name),
                "phone": None,
                "notes": None
            }
    
    for name, address in customer_data.items():
        if name not in customers:
            customers[name] = {
                "name": name,
                "address": address,
                "phone": None,
                "notes": None
            }
    
    print(f"\n   Unique farms: {len(farms)}")
    print(f"   Unique customers: {len(customers)}")
    
    print("\n   Connecting to Supabase...")
    try:
        supabase = get_supabase_client()
        print("   ✅ Connected successfully!")
    except Exception as e:
        print(f"   ❌ Error connecting to Supabase: {e}")
        sys.exit(1)
    
    farm_ids = insert_farms(supabase, farms)
    customer_ids = insert_customers(supabase, customers)
    inserted_count = insert_transactions(supabase, all_transactions, farm_ids, customer_ids)
    
    return supabase


def main():
    print("="*60)
    print("PIG FARM DATABASE DEPLOYMENT & DATA IMPORT")
    print("="*60)
    print(f"Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    try:
        # Step 1: Run migration
        run_migration_via_management_api()
        
        # Step 2: Import data
        supabase = import_data()
        
        # Step 3: Verify import
        stats = verify_import(supabase)
        
        print("\n" + "="*60)
        print("DEPLOYMENT COMPLETE")
        print("="*60)
        print(f"Finished at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        if stats:
            print(f"\n📈 Final Statistics:")
            print(f"   - Farms: {stats['farms']}")
            print(f"   - Customers: {stats['customers']}")
            print(f"   - Total Transactions: {stats['transactions']}")
            print(f"     * Heo thịt: {stats['heo_thit']}")
            print(f"     * Heo con: {stats['heo_con']}")
        
    except Exception as e:
        print(f"\n❌ Deployment failed: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
